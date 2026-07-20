"""Streamlit application for fetching EV Telemetry and DTC Data from LINKFMS API.

This file keeps the original app behavior (telemetry/DTC fetch, visualization, KPI)
but ensures:
- Login page is the first screen (no standalone username/password fields elsewhere)
- Login UI is isolated into agents/login_page_agent.py

NOTE: The app name/file name remains `linkfms_fetch_app.py` as requested.
"""

import io
import os
import streamlit as st
import pandas as pd
import sys
import time
import threading
from pathlib import Path
from datetime import date, timedelta
from time import perf_counter

import plotly.io as pio

# HTML export helpers (split out for customization)
from save_html_visualize_agent import build_visualize_export_html, VisualizeExportContext
from save_html_kpi_agent import build_kpi_export_html, KPIExportContext
from mail_html_agent import send_html_report_email, parse_email_list
from ollama_chat_agent import ask_ollama_fast, build_compact_context



sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parent.parent))

from linkfms_api_agent import LinkFMSAPIAgent, FETCH_SPEED_PRESETS
from linkfms_dtc_agent import LinkFMSDTCAgent
from sql_agent import SQLAgent, SQLConfig
from cache_fetch_agent import CacheFetchAgent
from data_sync_agent import DataSyncAgent
from sync_window_agent import SyncWindowAgent
from parameter_extraction_agent import ParameterExtractionAgent
from visualization_agent import VisualizationAgent
from kpi_agent_v2 import KPIAgent
from tmx_kpi_agent import TMXKPIAgent
from data_models.telemetry import TelemetryData, TelemetryParameter
from utils.login_auth import LoginManager

from login_page_agent import render_login_page

from save_html_visualize_agent import build_visualize_export_html, VisualizeExportContext
from save_html_kpi_agent import build_kpi_export_html, KPIExportContext



# "?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?
# PAGE CONFIG
# "?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?
st.set_page_config(page_title="🌐 TLD/XOPS VISUALIZATION TOOL", page_icon="T", layout="wide")

# "?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?
# INITIALIZE SESSION STATE
# "?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?
LoginManager.initialize_session()


def _record_timing(stage_timings, stage_name: str, stage_start: float) -> float:
    """Record a stage duration and print it to terminal logs."""
    elapsed = perf_counter() - stage_start
    stage_timings.append({"stage": stage_name, "seconds": round(elapsed, 3)})
    print(f"[TIMING] {stage_name}: {elapsed:.3f}s", flush=True)
    return perf_counter()


def _update_live_timing_ui(
    status_slot,
    table_slot,
    stage_timings,
    current_stage: str,
    is_complete: bool = False,
    active_stage: str | None = None,
    active_elapsed_seconds: float | None = None,
) -> None:
    """Render timing progress live in Streamlit while stages complete."""
    rows = list(stage_timings)
    if active_stage is not None and active_elapsed_seconds is not None:
        rows.append({"stage": f"{active_stage} (running)", "seconds": round(active_elapsed_seconds, 3)})

    timing_df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=["stage", "seconds"])
    table_slot.dataframe(timing_df, use_container_width=True, hide_index=True)

    if is_complete:
        status_slot.success(f"Timing complete. Final stage: {current_stage}")
    else:
        if active_stage is not None and active_elapsed_seconds is not None:
            status_slot.info(f"Running stage: {active_stage} | Elapsed: {int(active_elapsed_seconds)} sec")
        else:
            status_slot.info(f"Running stage: {current_stage}")


def _run_stage_with_live_timer(stage_name, status_slot, table_slot, stage_timings, stage_fn):
    """Run a blocking stage while updating elapsed seconds in realtime."""
    result_holder = {"value": None, "error": None}

    def _target():
        try:
            result_holder["value"] = stage_fn()
        except Exception as exc:
            result_holder["error"] = exc

    thread = threading.Thread(target=_target, daemon=True)
    stage_start = perf_counter()
    last_second = -1
    thread.start()

    while thread.is_alive():
        elapsed = perf_counter() - stage_start
        elapsed_sec = int(elapsed)
        if elapsed_sec != last_second:
            last_second = elapsed_sec
            _update_live_timing_ui(
                status_slot,
                table_slot,
                stage_timings,
                current_stage=stage_name,
                active_stage=stage_name,
                active_elapsed_seconds=elapsed,
            )
        time.sleep(0.1)

    thread.join()

    if result_holder["error"] is not None:
        raise result_holder["error"]

    elapsed = perf_counter() - stage_start
    stage_timings.append({"stage": stage_name, "seconds": round(elapsed, 3)})
    print(f"[TIMING] {stage_name}: {elapsed:.3f}s", flush=True)
    _update_live_timing_ui(status_slot, table_slot, stage_timings, current_stage=stage_name)
    return result_holder["value"]


def _build_default_sql_agent() -> SQLAgent:
    """Create SQL agent from environment defaults without exposing DB controls in UI."""
    return SQLAgent(
        SQLConfig(
            host=os.getenv("PGHOST", "localhost"),
            port=int(os.getenv("PGPORT", "5432")),
            database=os.getenv("PGDATABASE", "telemetry"),
            user=os.getenv("PGUSER", "postgres"),
            password=os.getenv("PGPASSWORD", ""),
            sslmode=os.getenv("PGSSLMODE", "prefer"),
        )
    )


@st.cache_resource(show_spinner=False)
def _start_server_live_sync_default() -> dict:
    """Start a shared live sync worker once per server process using env-based credentials."""
    plate_env = os.getenv("LINKFMS_SYNC_PLATE", "T118059").strip() or "T118059"
    poll_seconds = int(os.getenv("LINKFMS_SYNC_POLL_SECONDS", "30"))

    username = os.getenv("LINKFMS_SYNC_USERNAME", "")
    password = os.getenv("LINKFMS_SYNC_PASSWORD", "")
    if not username or not password:
        return {
            "running": False,
            "plate": plate_env,
            "reason": "Missing LINKFMS_SYNC_USERNAME or LINKFMS_SYNC_PASSWORD",
            "agent": None,
        }

    try:
        sql_agent = _build_default_sql_agent()
        sql_agent.initialize()

        api_agent = LinkFMSAPIAgent(username, password)
        sync_plates = [plate_env]
        plates_provider = None
        if plate_env.lower() in {"all", "everything", "*", "all_active", "all_active_vehicles"}:
            def _discover_active_plates() -> list[str]:
                today = date.today()
                discovered_plates = api_agent.get_available_vehicles(
                    start_date=(today - timedelta(days=1)).isoformat(),
                    end_date=today.isoformat(),
                )
                return [str(p).strip() for p in (discovered_plates or []) if str(p).strip()]

            plates_provider = _discover_active_plates
            discovered_now = plates_provider()
            if discovered_now:
                sync_plates = discovered_now

        sync_agent = DataSyncAgent(
            api_agent,
            sql_agent,
            poll_seconds=poll_seconds,
            speed_preset="fastest",
            initial_lookback_days=1,
            plates_provider=plates_provider,
            window_agent=SyncWindowAgent.from_env(),
        )
        sync_agent.start(sync_plates)

        display_plate = plate_env
        if len(sync_plates) > 1:
            display_plate = f"all_active_vehicles ({len(sync_plates)})"

        return {
            "running": True,
            "plate": display_plate,
            "reason": "",
            "agent": sync_agent,
        }
    except Exception as exc:
        return {
            "running": False,
            "plate": plate_env,
            "reason": str(exc),
            "agent": None,
        }


SERVER_LIVE_SYNC = _start_server_live_sync_default()


@st.fragment(run_every=2)
def _render_live_sync_status() -> None:
    """Render live sync status with periodic UI refresh."""
    sync_agent = SERVER_LIVE_SYNC.get("agent")
    sync_state = getattr(sync_agent, "state", None)
    sync_label = SERVER_LIVE_SYNC.get("plate", "T118059")
    current_plate = getattr(sync_state, "current_plate", None)

    if current_plate:
        st.caption(f"syncing.. {sync_label} | fetching now: {current_plate}")
    else:
        st.caption(f"syncing.. {sync_label}")

    active_plates = getattr(sync_state, "active_plates", None) or []
    if active_plates:
        with st.popover(f"Active sync plates ({len(active_plates)})"):
            current_plate = getattr(sync_state, "current_plate", None)
            for plate in active_plates:
                is_current = bool(current_plate and plate == current_plate)
                if is_current:
                    st.markdown(
                        f"<div style='padding:6px 10px;margin:4px 0;border-radius:8px;background:#0f172a;color:#ffffff;font-weight:700;border:1px solid #38bdf8;'>"
                        f"{plate} <span style='font-size:12px;opacity:0.9;'>(fetching now)</span>"
                        "</div>",
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(
                        f"<div style='padding:6px 10px;margin:4px 0;border-radius:8px;background:#f8fafc;color:#0f172a;border:1px solid #e5e7eb;'>"
                        f"{plate}"
                        "</div>",
                        unsafe_allow_html=True,
                    )


def show_main_app() -> None:
    """Main telemetry/DTC/KPI/visualization UI (rendered only after login)."""

    # Header with user info and logout
    col_title, col_logout = st.columns([4, 1])
    with col_title:
        st.title("🌐TLD/XOPS VISUALIZATION TOOL")
    with col_logout:
        st.markdown("")
        st.markdown("")
        if st.button("➜] Logout", use_container_width=True):
            st.session_state.authenticated = False
            st.session_state.user_info = None
            st.rerun()

    st.markdown(
        f"**Logged in as:** {st.session_state.user_info['name']} ({st.session_state.user_info['role']} 👤)"
    )
    _render_live_sync_status()
    st.markdown("---")

    st.subheader("Fetch Parameters")
    col1, col2, col3 = st.columns(3)
    with col1:
        plate_number = st.text_input("Vehicle Plate Number", placeholder="e.g., T118059")
    with col2:
        start_date = st.date_input("Start Date", value=date.today())
    with col3:
        end_date = st.date_input("End Date", value=date.today())

    TIMEZONE_OPTIONS = {
        "UTC": "UTC",
        "IST - India (UTC+5:30)": "Asia/Kolkata",
        "EST - USA Eastern (UTC-5)": "America/New_York",
        "CST - USA Central (UTC-6)": "America/Chicago",
        "PST - USA Pacific (UTC-8)": "America/Los_Angeles",
        "Paris - France (Europe/Paris, UTC+1/UTC+2 DST)": "Europe/Paris",
    }

    st.caption("Paris timezone is DST-aware. India is 3h30 ahead of Paris in summer, and 4h30 ahead in winter.")

    st.markdown("---")

    data_type = st.radio(
        "Select Data Type to Fetch",
options=[
            "Raw Telemetry",
            "DTC (Diagnostic Trouble Codes)",
            "Visualize",
            "Standard KPI Dashoard",
            "Advanced KPI Insights (TMX Style)",
            "AI Assistant",
        ],
        horizontal=True,
    )

    # "?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?
    # TELEMETRY
    # "?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?
    if data_type == "Raw Telemetry":
        col1, col2, col3 = st.columns(3)
        with col1:
            fetch_speed_preset = st.selectbox(
                "Fetch Speed Preset",
                options=list(FETCH_SPEED_PRESETS.keys()),
                index=list(FETCH_SPEED_PRESETS.keys()).index("fastest"),
                help="Controls page size, chunk size, pacing, cooldown, and retry behavior for telemetry fetches.",
            )
        with col2:
            alignment_method = st.selectbox(
                "Alignment Method",
                options=["forward_fill", "backward_fill", "interpolate", "nearest"],
                index=0,
            )
        with col3:
            selected_tz_label = st.selectbox("Display Timezone", options=list(TIMEZONE_OPTIONS.keys()), index=1)
        selected_tz = TIMEZONE_OPTIONS[selected_tz_label]

        preset_cfg = FETCH_SPEED_PRESETS[fetch_speed_preset]
        st.caption(
            f"Preset '{fetch_speed_preset}': page size {preset_cfg.page_size_telemetry:,}, chunk {preset_cfg.max_days_per_chunk} days, delay {preset_cfg.inter_request_delay_seconds:.1f}s, cooldown every {preset_cfg.cooldown_every_n_requests or 0} requests."
        )

        tracked_plate = plate_number.strip() if plate_number else ""
        if tracked_plate:
            st.caption(f"Live data fetch is ON | plate: {tracked_plate}")
        else:
            st.caption("Live data fetch is ON | plate: (enter plate number)")

        if st.button("Fetch Telemetry Data", type="primary"):
            if not plate_number:
                st.error("Please enter a vehicle plate number.")
                return

            try:
                total_start = perf_counter()
                stage_timings = []
                live_timing_status = st.empty()
                live_timing_table = st.empty()
                _update_live_timing_ui(
                    live_timing_status,
                    live_timing_table,
                    stage_timings,
                    current_stage="initializing",
                )

                creds = st.session_state.get("user_info") or {}
                api_agent = LinkFMSAPIAgent(
                    creds.get("username") or creds.get("userid") or "",
                    creds.get("password") or "",
                )

                sql_agent = _build_default_sql_agent()
                sql_agent.initialize()
                cache_agent = CacheFetchAgent(sql_agent=sql_agent, api_agent=api_agent)

                def _stage_api_fetch():
                    return cache_agent.fetch_db_first(
                        plate_number=plate_number,
                        start_date=start_date.isoformat(),
                        end_date=end_date.isoformat(),
                        speed_preset=fetch_speed_preset,
                    )

                batch = _run_stage_with_live_timer(
                    "api_fetch", live_timing_status, live_timing_table, stage_timings, _stage_api_fetch
                )

                st.caption(
                    f"Data source: {batch.source} | DB rows: {batch.cached_rows} | Newly fetched rows: {batch.fetched_rows}"
                )

                if not batch or batch.dataframe.empty:
                    st.warning("No data found.")
                    total_elapsed = perf_counter() - total_start
                    stage_timings.append({"stage": "total_pipeline", "seconds": round(total_elapsed, 3)})
                    print(f"[TIMING] total_pipeline: {total_elapsed:.3f}s", flush=True)
                    st.session_state["telemetry_timing_df"] = pd.DataFrame(stage_timings)
                    _update_live_timing_ui(
                        live_timing_status,
                        live_timing_table,
                        stage_timings,
                        current_stage="total_pipeline",
                        is_complete=True,
                    )
                    return

                timestamp_col_name = "timestamp"

                def _stage_prepare_dataframe():
                    combined_df_local = batch.dataframe.copy()
                    if timestamp_col_name in combined_df_local.columns:
                        combined_df_local[timestamp_col_name] = pd.to_datetime(
                            combined_df_local[timestamp_col_name], errors="coerce", utc=True
                        )
                        combined_df_local = combined_df_local.sort_values(timestamp_col_name).reset_index(drop=True)

                    telemetry_data_local = TelemetryData(
                        dataframe=combined_df_local,
                        parameters=[
                            TelemetryParameter(name=col)
                            for col in combined_df_local.columns
                            if col != timestamp_col_name
                        ],
                        timestamp_column=timestamp_col_name,
                        source_file=f"LINKFMS API - {plate_number}",
                    )
                    return telemetry_data_local

                telemetry_data = _run_stage_with_live_timer(
                    "prepare_dataframe", live_timing_status, live_timing_table, stage_timings, _stage_prepare_dataframe
                )

                def _stage_parameter_extraction():
                    extraction_agent = ParameterExtractionAgent()
                    extraction_result = extraction_agent.extract_parameters(
                        telemetry_data.dataframe, source_column="telemetry_raw"
                    )

                    raw_api_cols_to_drop = [
                        "telemetry_raw",
                        "accuracy",
                        "gpsProvider",
                        "__typename",
                        "longitude",
                        "latitude",
                        "id",
                        "plateNumber",
                        "VehicleEntity Max RPM",
                        "VehicleEntity Max Idle",
                        "VehicleEntity Max Speed",
                        "EV Charger Output Current",
                        "Veh Ignition",
                        "Veh Inching",
                        "EV Battery Cell Temp Min",
                        "EV Battery Cell Temp Max",
                        "EV Battery Cell Temp Min (C)",
                        "EV Battery Cell Temp Max (C)",
                        "OBU Vin State",
                        "OBU Vin Data Mb/M",
                        "OBU Internal Temperature (C)",
                        "Alrm Harsh Braking",
                        "Alrm Excessive Acceleration",
                        "Alrm No Trip Motion",
                        "VCM Vehicle Access Control State I/O",
                    ]

                    extracted_df = extraction_result.dataframe.drop(
                        columns=[c for c in raw_api_cols_to_drop if c in extraction_result.dataframe.columns],
                        errors="ignore",
                    )
                    extracted_df = extracted_df.loc[:, ~extracted_df.columns.duplicated(keep="first")]

                    # Keep engine code related fields as true observed values only.
                    # Missing values should stay missing (not string placeholders).
                    for col in ["EngineCodeDescription", "Source", "EngineCode"]:
                        if col in extracted_df.columns:
                            extracted_df[col] = extracted_df[col].replace(
                                {"None": pd.NA, "nan": pd.NA, "NaN": pd.NA, "": pd.NA}
                            )
                    return extracted_df

                extracted_df = _run_stage_with_live_timer(
                    "parameter_extraction", live_timing_status, live_timing_table, stage_timings, _stage_parameter_extraction
                )

                def _stage_alignment_fill():
                    local_df = extracted_df.copy()
                    local_df = local_df.dropna(subset=[timestamp_col_name])
                    local_df = local_df.sort_values(timestamp_col_name).reset_index(drop=True)

                    non_ts_cols = [c for c in local_df.columns if c != timestamp_col_name]
                    no_fill_cols = [
                        c for c in ["EngineCodeDescription", "Source", "EngineCode"]
                        if c in non_ts_cols
                    ]
                    fill_cols = [c for c in non_ts_cols if c not in no_fill_cols]

                    # Track which values are actually forward/back-filled.
                    # For each column: mark True where the value was NaN in the original extracted_df,
                    # but becomes non-NaN after filling.
                    original_is_nan = local_df[non_ts_cols].isna()
                    filled_mask = pd.DataFrame(False, index=local_df.index, columns=non_ts_cols)

                    if alignment_method == "forward_fill":
                        if fill_cols:
                            local_df[fill_cols] = local_df[fill_cols].ffill()
                            filled_mask.loc[:, fill_cols] = original_is_nan[fill_cols] & local_df[fill_cols].notna()
                    elif alignment_method == "backward_fill":
                        if fill_cols:
                            local_df[fill_cols] = local_df[fill_cols].bfill()
                            filled_mask.loc[:, fill_cols] = original_is_nan[fill_cols] & local_df[fill_cols].notna()
                    elif alignment_method in ("interpolate", "nearest"):
                        numeric_cols = local_df[fill_cols].select_dtypes(include="number").columns.tolist() if fill_cols else []
                        if numeric_cols:
                            local_df[numeric_cols] = local_df[numeric_cols].interpolate(
                                method="linear", limit_direction="both"
                            )
                        # Remaining NaNs are handled with ffill/bfill
                        if fill_cols:
                            local_df[fill_cols] = local_df[fill_cols].ffill().bfill()
                            filled_mask.loc[:, fill_cols] = original_is_nan[fill_cols] & local_df[fill_cols].notna()
                    else:
                        if fill_cols:
                            local_df[fill_cols] = local_df[fill_cols].ffill().bfill()
                            filled_mask.loc[:, fill_cols] = original_is_nan[fill_cols] & local_df[fill_cols].notna()

                    display_df_local = local_df.rename(columns={"timestamp": "dateProcessed"})

                    # Store fill masks for UI highlighting (keyed by column name).
                    fill_masks = {}
                    for col in non_ts_cols:
                        if col in filled_mask.columns:
                            # Align indices after rename (same as extracted_df)
                            fill_masks[col] = filled_mask[col].tolist()
                    return display_df_local, fill_masks

                display_df, fill_masks = _run_stage_with_live_timer(
                    "alignment_fill", live_timing_status, live_timing_table, stage_timings, _stage_alignment_fill
                )
                st.session_state['fill_masks'] = fill_masks


                # timezone conversion for any date columns
                def _stage_timezone_conversion():
                    local_df = display_df.copy()
                    date_cols = [c for c in ["date", "dateReceived", "dateProcessed"] if c in local_df.columns]
                    for col in date_cols:
                        local_df[col] = (
                            pd.to_datetime(local_df[col], errors="coerce", utc=True)
                            .dt.tz_convert(selected_tz)
                            .dt.tz_localize(None)
                        )
                    return local_df

                display_df = _run_stage_with_live_timer(
                    "timezone_conversion", live_timing_status, live_timing_table, stage_timings, _stage_timezone_conversion
                )

                st.session_state["display_df"] = display_df
                st.session_state["tele_plate"] = plate_number
                st.session_state["tele_start"] = start_date
                st.session_state["tele_end"] = end_date
                st.session_state["tele_tz"] = selected_tz_label

                total_elapsed = perf_counter() - total_start
                stage_timings.append({"stage": "total_pipeline", "seconds": round(total_elapsed, 3)})
                print(f"[TIMING] total_pipeline: {total_elapsed:.3f}s", flush=True)
                st.session_state["telemetry_timing_df"] = pd.DataFrame(stage_timings)
                _update_live_timing_ui(
                    live_timing_status,
                    live_timing_table,
                    stage_timings,
                    current_stage="total_pipeline",
                    is_complete=True,
                )

                st.success(f"{len(display_df)} records fetched for {plate_number}. TZ: {selected_tz_label}")

            except Exception as e:
                st.error(f"An unexpected error occurred: {e}")

        # Persistent preview
        if "display_df" in st.session_state:
            display_df = st.session_state["display_df"]
            st.subheader("Fetched Data Preview (Bold = forward/back-filled)")

            # Keep internal datetimes for analytics, but render/export with seconds.
            display_df_out = display_df.copy()
            out_date_cols = [c for c in ["date", "dateReceived", "dateProcessed", "timestamp"] if c in display_df_out.columns]
            for col in out_date_cols:
                display_df_out[col] = pd.to_datetime(display_df_out[col], errors="coerce").dt.strftime("%Y-%m-%d %H:%M:%S")

            # If we have fill masks from alignment stage, render styled table:
            # - Bold: value was filled (original NaN ?' non-NaN)
            # - Normal: original generated value
            fill_masks = st.session_state.get("fill_masks") or {}
            # Avoid Streamlit/Pandas Styler hard limits by only styling a preview subset.
            # (Styler render cell limit is commonly 262144 cells by default.)
            preview_rows = 500
            hidden_preview_cols = ["Speed (km/h)"]
            display_preview_df = display_df_out.drop(columns=hidden_preview_cols, errors="ignore")
            display_preview = display_preview_df.head(preview_rows)
            if fill_masks:
                styled = display_preview.style

                for col, mask_list in fill_masks.items():
                    if col in display_preview.columns:
                        mask_series = pd.Series(mask_list, index=display_df.index).astype(bool)
                        styled = styled.apply(
                            lambda s, ms=mask_series: ["font-weight: 700" if ms.iloc[i] else ""
                                                       for i in range(len(s))],
                            subset=[col],
                        )


                st.dataframe(styled, use_container_width=True)
            else:
                st.dataframe(display_preview_df, use_container_width=True)


            date_col_display = "dateProcessed" if "dateProcessed" in display_df.columns else "timestamp"
            st.write(f"**Total Records:** {len(display_df)}")
            st.write(
                f"**Date Range:** {display_df[date_col_display].min()} to {display_df[date_col_display].max()}"
            )
            st.write(
                f"**Memory Usage:** {display_df.memory_usage(deep=True).sum() / 1024 / 1024:.2f} MB"
            )

            timing_df = st.session_state.get("telemetry_timing_df")
            if isinstance(timing_df, pd.DataFrame) and not timing_df.empty:
                with st.expander("Performance timings (seconds)", expanded=True):
                    st.dataframe(timing_df, use_container_width=True, hide_index=True)
                    total_match = timing_df[timing_df["stage"] == "total_pipeline"]
                    if not total_match.empty:
                        st.caption(f"Total pipeline time: {float(total_match.iloc[-1]['seconds']):.3f}s")

            st.markdown("---")
            # For downloads, also preserve the filled/non-filled info by adding a column-wise mask.
            # Excel export then uses that mask to apply real bold formatting.
            csv_df = display_df_out.copy()
            fill_masks = st.session_state.get("fill_masks") or {}

            for col, mask_list in fill_masks.items():
                if col in csv_df.columns:
                    csv_df[f"{col}__is_filled"] = mask_list

            col_csv, col_excel = st.columns(2)
            with col_csv:
                st.download_button(
                    label="Download as CSV",
                    data=csv_df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig"),
                    file_name=f"telemetry_{st.session_state['tele_plate']}_{st.session_state['tele_start']}_to_{st.session_state['tele_end']}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
            with col_excel:
                # Excel formatting: bold cells for values that were actually filled.
                import openpyxl
                from openpyxl.styles import Font

                excel_buf = io.BytesIO()
                with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
                    csv_df.to_excel(writer, index=False, sheet_name="telemetry")

                excel_buf.seek(0)
                wb = openpyxl.load_workbook(excel_buf)
                ws = wb["telemetry"]

                bold_font = Font(bold=True)

                # Map header -> column index
                header = [c.value for c in ws[1]]
                header_to_idx = {name: i for i, name in enumerate(header) if name is not None}

                # For each original parameter column, look for its __is_filled column to decide bolding.
                # IMPORTANT: we iterate by row count from the Excel sheet, not from the mask length,
                # to avoid corrupting the workbook when the mask length != exported rows.
                max_excel_rows = ws.max_row - 1  # exclude header

                for col, mask_list in (fill_masks or {}).items():
                    if col not in header_to_idx:
                        continue

                    is_filled_col_name = f"{col}__is_filled"
                    if is_filled_col_name not in header_to_idx:
                        continue

                    col_idx = header_to_idx[col]  # 0-based

                    # Rows in openpyxl are 1-based; header is row 1.
                    mask_len = len(mask_list)
                    rows_to_apply = min(max_excel_rows, mask_len)

                    for r in range(rows_to_apply):
                        excel_row = 2 + r
                        if mask_list[r]:
                            ws.cell(row=excel_row, column=1 + col_idx).font = bold_font


                wb.save(excel_buf)
                excel_buf.seek(0)

                st.download_button(
                    label="Download as Excel",
                    data=excel_buf,
                    file_name=f"telemetry_{st.session_state['tele_plate']}_{st.session_state['tele_start']}_to_{st.session_state['tele_end']}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )


    # "?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?
    # DTC
    # "?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?
    elif data_type == "DTC (Diagnostic Trouble Codes)":
        selected_tz_label_dtc = st.selectbox(
            "Display Timezone", options=list(TIMEZONE_OPTIONS.keys()), index=1, key="tz_dtc"
        )
        selected_tz_dtc = TIMEZONE_OPTIONS[selected_tz_label_dtc]

        if st.button("Fetch DTC Data", type="primary"):
            if not plate_number:
                st.error("Please enter a vehicle plate number.")
                return

            try:
                # DTC agent auth uses username/password from LoginManager.
                user_info = st.session_state.get("user_info") or {}
                linkfms_username = user_info.get("username") or user_info.get("userid") or ""
                linkfms_password = user_info.get("password") or ""

                dtc_agent = LinkFMSDTCAgent(linkfms_username, linkfms_password)


                with st.spinner("Fetching DTC records..."):
                    dtc_df = dtc_agent.fetch(
                        plate_number=plate_number,
                        start_date=start_date.isoformat(),
                        end_date=end_date.isoformat(),
                    )

                if dtc_df is not None and not dtc_df.empty:
                    dtc_df["timestamp"] = (
                        dtc_df["timestamp"].dt.tz_convert(selected_tz_dtc).dt.tz_localize(None)
                    )
                    st.session_state["dtc_df"] = dtc_df
                    st.session_state["dtc_plate"] = plate_number
                    st.session_state["dtc_start"] = start_date
                    st.session_state["dtc_end"] = end_date
                    st.success(f"Fetched {len(dtc_df)} DTC records for {plate_number}.")
                else:
                    st.warning("No DTC records found.")

            except Exception as e:
                st.error(f"An unexpected error occurred: {e}")

        if "dtc_df" in st.session_state:
            dtc_df = st.session_state["dtc_df"]
            st.subheader("DTC Data Preview")
            st.dataframe(dtc_df, use_container_width=True)

            st.write(f"**Total DTC Records:** {len(dtc_df)}")
            if "code" in dtc_df.columns:
                st.write(f"**Unique Codes:** {dtc_df['code'].nunique()}")
            if "timestamp" in dtc_df.columns:
                st.write(f"**Date Range:** {dtc_df['timestamp'].min()} to {dtc_df['timestamp'].max()}")

            st.markdown("---")
            dtc_csv_df = dtc_df.copy().where(dtc_df.notna(), other="")

            col_csv2, col_excel2 = st.columns(2)
            with col_csv2:
                st.download_button(
                    label="Download DTC as CSV",
                    data=dtc_csv_df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig"),
                    file_name=f"dtc_{st.session_state['dtc_plate']}_{st.session_state['dtc_start']}_to_{st.session_state['dtc_end']}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
            with col_excel2:
                dtc_buf = io.BytesIO()
                dtc_csv_df.to_excel(dtc_buf, index=False, engine="openpyxl")
                dtc_buf.seek(0)
                st.download_button(
                    label="Download DTC as Excel",
                    data=dtc_buf,
                    file_name=f"dtc_{st.session_state['dtc_plate']}_{st.session_state['dtc_start']}_to_{st.session_state['dtc_end']}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

    # "?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?
    # VISUALIZE
    # "?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?
    elif data_type == "Visualize":
        viz = VisualizationAgent()

        # Keep reference(s) to generated charts so we can export them as interactive HTML.
        last_generated_fig = None
        generated_figs_for_html = []

        has_telemetry = "display_df" in st.session_state
        has_dtc = "dtc_df" in st.session_state

        if not has_telemetry and not has_dtc:
            st.info("Fetch Telemetry or DTC data first, then come here to visualize.")
            return

        viz_section = st.radio(
            "Visualize",
            options=(["Telemetry Charts"] if has_telemetry else []) + (["DTC Charts"] if has_dtc else []),
            horizontal=True,
        )

        if viz_section == "Telemetry Charts":
            df = st.session_state["display_df"]
            timestamp_col = "dateProcessed" if "dateProcessed" in df.columns else df.columns[0]
            numeric_cols = df.select_dtypes(include="number").columns.tolist()

            chart_category = st.selectbox(
                "Select Chart Category",
                ["Key Telemetry Charts", "Statistical & Relationship Charts", "Custom Chart Builder"],
            )

            if chart_category == "Key Telemetry Charts":
                chart_type = st.selectbox(
                    "Select Chart",
                    [
                        "SOC & Battery Current (Dual Axis)",
                        "Temperature Parameters",
                        "Speed Over Time",
                        "SOC Daily Min/Max",
                        "Charging Sessions",
                        "Custom Time Series",
                    ],
                )

                if chart_type == "SOC & Battery Current (Dual Axis)":
                    fig = viz.soc_current_dual(df, timestamp_col)
                    last_generated_fig = fig
                    st.plotly_chart(fig, use_container_width=True)
                elif chart_type == "Temperature Parameters":
                    fig = viz.temperature_chart(df, timestamp_col)
                    last_generated_fig = fig
                    st.plotly_chart(fig, use_container_width=True)
                elif chart_type == "Speed Over Time":
                    speed_col = st.selectbox(
                        "Speed column", [c for c in df.columns if "speed" in c.lower()] or numeric_cols
                    )
                    fig = viz.speed_chart(df, timestamp_col, speed_col)
                    last_generated_fig = fig
                    st.plotly_chart(fig, use_container_width=True)
                elif chart_type == "SOC Daily Min/Max":
                    soc_col = st.selectbox(
                        "SOC column",
                        [
                            c
                            for c in df.columns
                            if ("soc" in c.lower() or "charge" in c.lower() or "energy" in c.lower())
                        ]
                        or numeric_cols,
                    )
                    fig = viz.soc_daily_bar(df, timestamp_col, soc_col)
                    last_generated_fig = fig
                    st.plotly_chart(fig, use_container_width=True)
                elif chart_type == "Charging Sessions":
                    soc_col = st.selectbox(
                        "SOC column",
                        [
                            c
                            for c in df.columns
                            if ("soc" in c.lower() or "charge" in c.lower() or "energy" in c.lower())
                        ]
                        or numeric_cols,
                    )
                    charger_col = st.selectbox(
                        "Charger State column", [c for c in df.columns if "charger" in c.lower()] or [""],
                    )
                    fig = viz.charging_sessions(df, timestamp_col, soc_col, charger_col)
                    last_generated_fig = fig
                    st.plotly_chart(fig, use_container_width=True)
                elif chart_type == "Custom Time Series":
                    selected_params = st.multiselect("Select parameters", numeric_cols, default=numeric_cols[:3])
                    if selected_params:
                        fig = viz.time_series(df, timestamp_col, selected_params)
                        last_generated_fig = fig
                        st.plotly_chart(fig, use_container_width=True)

            elif chart_category == "Statistical & Relationship Charts":
                chart_type = st.selectbox("Select Chart", ["Scatter Correlation", "Correlation Heatmap", "Box Plots", "Histogram"])

                if chart_type == "Scatter Correlation":
                    col1, col2 = st.columns(2)
                    with col1:
                        x_col = st.selectbox("X axis", numeric_cols)
                    with col2:
                        y_col = st.selectbox("Y axis", numeric_cols, index=min(1, len(numeric_cols) - 1))
                    st.plotly_chart(viz.scatter_correlation(df, x_col, y_col), use_container_width=True)
                elif chart_type == "Correlation Heatmap":
                    st.plotly_chart(viz.heatmap_correlation(df), use_container_width=True)
                elif chart_type == "Box Plots":
                    selected_params = st.multiselect("Select parameters", numeric_cols, default=numeric_cols[:5])
                    if selected_params:
                        st.plotly_chart(viz.box_plots(df, selected_params), use_container_width=True)
                elif chart_type == "Histogram":
                    param = st.selectbox("Select parameter", numeric_cols)
                    bins = st.slider("Bins", 10, 100, 30)
                    st.plotly_chart(viz.histogram(df, param, bins), use_container_width=True)

            elif chart_category == "Custom Chart Builder":
                all_cols = [c for c in df.columns if c != timestamp_col]

                if "tele_custom_charts" not in st.session_state:
                    st.session_state.tele_custom_charts = []

                st.markdown("#### Custom Chart Builder")
                st.markdown("**Create New Chart**")

                col1, col2, col3 = st.columns(3)
                with col1:
                    x_col = st.selectbox("X Axis", [timestamp_col] + all_cols, key="tele_x_col")
                    graph_type = st.selectbox(
                        "Chart Type",
                        ["line", "bar", "scatter", "area", "box", "violin", "histogram", "pie", "pareto"],
                        key="tele_graph_type",
                    )
                with col2:
                    chart_title = st.text_input("Chart Title", value="Custom Chart", key="tele_chart_title")
                    y_cols = st.multiselect(
                        "Y Axis (select one or more)",
                        all_cols,
                        default=numeric_cols[:1] if numeric_cols else all_cols[:1],
                        key="tele_y_cols",
                    )
                with col3:
                    add_chart_btn = st.button("Add Chart", key="tele_add_chart_btn", use_container_width=True)

                colors = []
                if y_cols:
                    st.markdown("**Colors** (one per Y column)")
                    color_cols = st.columns(min(len(y_cols), 4))
                    default_hex = ["#636efa", "#ef553b", "#00cc96", "#ab63fa", "#ffa15a", "#19d3f3"]
                    for i, col in enumerate(y_cols):
                        with color_cols[i % len(color_cols)]:
                            colors.append(
                                st.color_picker(col, value=default_hex[i % len(default_hex)], key=f"tele_color_{i}")
                            )

                if add_chart_btn and y_cols:
                    st.session_state.tele_custom_charts.append(
                        {
                            "x_col": x_col,
                            "y_cols": y_cols,
                            "graph_type": graph_type,
                            "colors": colors,
                            "title": chart_title,
                        }
                    )
                    st.rerun()

                # Reset export figs for this custom builder render pass
                generated_figs_for_html = []

                if st.session_state.tele_custom_charts:
                    st.markdown("---")
                    st.markdown("#### Generated Charts")
                    for idx, chart_cfg in enumerate(st.session_state.tele_custom_charts):
                        col_chart, col_delete = st.columns([0.95, 0.05])
                        with col_chart:
                            fig = viz.custom_chart(
                                df,
                                chart_cfg["x_col"],
                                chart_cfg["y_cols"],
                                chart_cfg["graph_type"],
                                chart_cfg["colors"],
                                chart_cfg["title"],
                            )
                            last_generated_fig = fig
                            generated_figs_for_html.append(fig)
                            st.plotly_chart(fig, use_container_width=True)
                        with col_delete:
                            if st.button("🗑️", key=f"tele_delete_{idx}"):
                                st.session_state.tele_custom_charts.pop(idx)
                                st.rerun()

        else:  # DTC charts
            dtc_df = st.session_state["dtc_df"]

            chart_category_dtc = st.selectbox("Select DTC Chart Category", ["Key DTC Charts", "Custom Chart Builder"])

            if chart_category_dtc == "Key DTC Charts":
                chart_type = st.selectbox(
                    "Select Chart",
                    ["DTC Frequency Bar", "DTC Event Timeline", "DTC Severity Breakdown", "DTC Unique Codes Over Time", "Daily DTC Count"],
                )

                if chart_type == "DTC Frequency Bar":
                    fig = viz.dtc_frequency_bar(dtc_df)
                    last_generated_fig = fig
                    st.plotly_chart(fig, use_container_width=True)
                elif chart_type == "DTC Event Timeline":
                    fig = viz.dtc_timeline(dtc_df)
                    last_generated_fig = fig
                    st.plotly_chart(fig, use_container_width=True)
                elif chart_type == "DTC Severity Breakdown":
                    fig = viz.dtc_severity_pie(dtc_df)
                    last_generated_fig = fig
                    st.plotly_chart(fig, use_container_width=True)
                elif chart_type == "DTC Unique Codes Over Time":
                    fig = viz.dtc_unique_codes_over_time(dtc_df)
                    last_generated_fig = fig
                    st.plotly_chart(fig, use_container_width=True)
                elif chart_type == "Daily DTC Count":
                    fig = viz.dtc_daily_count(dtc_df)
                    last_generated_fig = fig
                    st.plotly_chart(fig, use_container_width=True)

            else:
                all_dtc_cols = dtc_df.columns.tolist()
                numeric_dtc_cols = dtc_df.select_dtypes(include=["number"]).columns.tolist()
                available_for_axis = [c for c in all_dtc_cols if c != "timestamp"]

                if "dtc_custom_charts" not in st.session_state:
                    st.session_state.dtc_custom_charts = []

                st.markdown("#### Custom DTC Chart Builder")
                st.markdown("**Create New Chart**")

                col1, col2, col3 = st.columns(3)
                with col1:
                    dtc_x_col = st.selectbox("X Axis", ["timestamp"] + available_for_axis, key="dtc_x_col_select")
                    dtc_graph_type = st.selectbox(
                        "Chart Type",
                        ["line", "bar", "scatter", "area", "box", "violin", "histogram", "pie", "pareto"],
                        key="dtc_graph_type_select",
                    )
                with col2:
                    dtc_chart_title = st.text_input("Chart Title", value="Custom DTC Chart", key="dtc_chart_title_input")
                    y_axis_type = st.radio("Y Axis Type", ["Numeric Only", "All Columns"], horizontal=True, key="dtc_y_type_select")
                with col3:
                    dtc_add_chart_btn = st.button("z. Add Chart", key="dtc_add_chart_btn", use_container_width=True)

                dtc_y_options = numeric_dtc_cols if y_axis_type == "Numeric Only" else available_for_axis
                dtc_y_cols = st.multiselect(
                    "Y Axis Columns (select one or more)",
                    dtc_y_options,
                    default=dtc_y_options[:1] if dtc_y_options else [],
                    key="dtc_y_cols_select",
                )

                dtc_colors = []
                if dtc_y_cols:
                    st.markdown("**Colors** (one per Y column)")
                    color_cols_count = min(len(dtc_y_cols), 4)
                    dtc_color_cols = st.columns(color_cols_count)
                    default_hex = ["#636efa", "#ef553b", "#00cc96", "#ab63fa", "#ffa15a", "#19d3f3"]
                    for i, col in enumerate(dtc_y_cols):
                        with dtc_color_cols[i % color_cols_count]:
                            dtc_colors.append(
                                st.color_picker(col, value=default_hex[i % len(default_hex)], key=f"dtc_color_{i}_{col}")
                            )

                if dtc_add_chart_btn and dtc_y_cols:
                    st.session_state.dtc_custom_charts.append(
                        {
                            "x_col": dtc_x_col,
                            "y_cols": dtc_y_cols,
                            "graph_type": dtc_graph_type,
                            "colors": dtc_colors,
                            "title": dtc_chart_title,
                        }
                    )
                    st.rerun()

                # Reset export figs for this custom builder render pass
                generated_figs_for_html = []

                if st.session_state.dtc_custom_charts:
                    st.markdown("---")
                    st.markdown("#### Generated Charts")
                    for idx, chart_cfg in enumerate(st.session_state.dtc_custom_charts):
                        col_chart, col_delete = st.columns([0.95, 0.05])
                        with col_chart:
                            fig = viz.custom_chart(
                                dtc_df,
                                chart_cfg["x_col"],
                                chart_cfg["y_cols"],
                                chart_cfg["graph_type"],
                                chart_cfg["colors"],
                                chart_cfg["title"],
                            )
                            last_generated_fig = fig
                            generated_figs_for_html.append(fig)
                            st.plotly_chart(fig, use_container_width=True)
                        with col_delete:
                            if st.button("🗑️", key=f"dtc_delete_{idx}"):
                                st.session_state.dtc_custom_charts.pop(idx)
                                st.rerun()

        # If a chart was generated, allow exporting it as interactive HTML.
        if last_generated_fig is not None:
            figs_to_export = []
            if generated_figs_for_html:
                figs_to_export = generated_figs_for_html
            else:
                figs_to_export = [last_generated_fig]

            if figs_to_export:
                html_title = f"Custom Charts Export ({plate_number if 'plate_number' in locals() else 'Session'})"
                charts_html = "".join(
                    [
                        pio.to_html(
                            fig,
                            full_html=False,
                            include_plotlyjs=("cdn" if i == 0 else False),
                        )
                        for i, fig in enumerate(figs_to_export)
                    ]
                )
                export_html = (
                    "<!DOCTYPE html>\n"
                    "<html lang='en'>\n"
                    "<head>\n"
                    "<meta charset='utf-8'/>\n"
                    "<meta name='viewport' content='width=device-width, initial-scale=1'/>\n"
                    f"<title>{html_title}</title>\n"
                    "</head>\n"
                    "<body style='margin:0;padding:16px;background:#0b1220;color:white;font-family:Segoe UI,Arial,sans-serif;'>\n"
                    f"<h2 style='margin:0 0 14px 0;font-size:20px;'>{html_title}</h2>\n"
                    "<div style='display:grid;gap:18px;'>\n"
                    f"{charts_html}\n"
                    "</div>\n"
                    "</body>\n"
                    "</html>"
                )

                st.download_button(
                    label="Save Generated Graphs as HTML",
                    data=export_html,
                    file_name=("custom_charts_export.html"),
                    mime="text/html",
                    use_container_width=True,
                )



    # "?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?
    # KPI
    # "?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?
    elif data_type == "Standard KPI Dashoard":
        if "display_df" not in st.session_state:
            st.info("Fetch telemetry data first, then come here to generate KPI dashboard.")
            return

        generated_figs = []

        df = st.session_state["display_df"]
        plate_number = st.session_state.get("tele_plate", "Unknown")
        start_date = st.session_state.get("tele_start", "N/A")
        end_date = st.session_state.get("tele_end", "N/A")
        selected_tz = st.session_state.get("tele_tz", "UTC")

        st.markdown(
            f"Vehicle Performance Report\n**Plate:** {plate_number} | **Period:** {start_date} to {end_date} | **TZ:** {selected_tz}"
        )
        st.markdown("---")

        kpi = KPIAgent()
        timestamp_col = "dateProcessed" if "dateProcessed" in df.columns else "timestamp"

        with st.spinner("Calculating KPI metrics..."):
            metrics = kpi.calculate_metrics(
                df,
                timestamp_col=timestamp_col,
                engine_state_col="engineState" if "engineState" in df.columns else None,
                motor_hour_col="motorHour" if "motorHour" in df.columns else None,
                speed_col="speed" if "speed" in df.columns else None,
                odometer_col="odometer" if "odometer" in df.columns else None,
            )

        st.subheader("Overall KPI Metrics")
        with st.expander("View Formulas"):
            st.markdown("**Formulas Used:**")
            for formula_name, formula_text in metrics.formulas.items():
                label = formula_name.replace("_", " ").title()
                st.markdown(f"- **{label}** = {formula_text}")

        stats = metrics.overall_stats
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Running %", f"{stats['running_pct']:.1f}%", help="Percentage of time vehicle was running")
        with col2:
            st.metric("Idle %", f"{stats['idle_pct']:.1f}%", help="Percentage of time vehicle was idle")
        with col3:
            st.metric("Stopped %", f"{stats['stopped_pct']:.1f}%", help="Percentage of time vehicle was stopped")
        with col4:
            st.metric("Total Distance", f"{stats['total_distance']:.1f} km", help="Total distance traveled")

        col5, col6, col7, col8 = st.columns(4)
        with col5:
            st.metric("Running Hrs", f"{stats['total_running_hours']:.1f} h", help="Total running time")
        with col6:
            st.metric("Idle Hrs", f"{stats['total_idle_hours']:.1f} h", help="Total idle time")
        with col7:
            st.metric("Motor Hrs", f"{stats['total_motor_hours']:.1f} h", help="Total positive motor hour delta")
        with col8:
            st.metric("Active %", f"{stats['active_pct']:.1f}%", help="Running + idle percentage")

        col9, col10, col11, col12 = st.columns(4)
        with col9:
            st.metric("Avg Speed", f"{stats['avg_speed']:.1f} km/h", help="Weighted average speed across all time")
        with col10:
            st.metric("Avg Moving Speed", f"{stats['avg_moving_speed']:.1f} km/h", help="Average speed only while running")
        with col11:
            st.metric("Dist / Running Hr", f"{stats['distance_per_running_hour']:.1f} km/h", help="Distance efficiency over running hours")
        with col12:
            st.metric("Dist / Motor Hr", f"{stats['distance_per_motor_hour']:.1f} km/h", help="Distance efficiency over motor hours")

        insight_col1, insight_col2, insight_col3, insight_col4 = st.columns(4)
        with insight_col1:
            st.metric("Max Speed", f"{stats['max_speed']:.1f} km/h")
        with insight_col2:
            st.metric("Avg Daily Distance", f"{stats['avg_daily_distance']:.1f} km")
        with insight_col3:
            st.metric("Peak Daily Distance", f"{stats['peak_daily_distance']:.1f} km")
        with insight_col4:
            st.metric("Days With Data", f"{stats['days_with_data']}")

        st.markdown("---")

        st.subheader("Daily Metrics and Odometer")
        tab1, tab2 = st.tabs(["Daily Hours & Percentages", "Daily Odometer"])

        with tab1:
            col_chart1, col_chart2 = st.columns(2)
            with col_chart1:
                fig1 = kpi.create_daily_hours_chart(metrics.daily_metrics)
                generated_figs.append(fig1)
                st.plotly_chart(fig1, use_container_width=True)
            with col_chart2:
                fig2 = kpi.create_daily_percentage_chart(metrics.daily_metrics)
                generated_figs.append(fig2)
                st.plotly_chart(fig2, use_container_width=True)

            daily_display = metrics.daily_metrics[
                [
                    "date",
                    "running_hours",
                    "idle_hours",
                    "stopped_hours",
                    "active_hours",
                    "motor_hours",
                    "distance_km",
                    "running_pct",
                    "idle_pct",
                    "stopped_pct",
                    "active_pct",
                    "avg_speed",
                    "avg_moving_speed",
                    "max_speed",
                    "distance_per_running_hour",
                    "distance_per_motor_hour",
                    "idle_to_running_ratio",
                ]
            ].copy()
            daily_display.columns = [
                "Date",
                "Running Hrs",
                "Idle Hrs",
                "Stopped Hrs",
                "Active Hrs",
                "Motor Hrs",
                "Distance (km)",
                "Running %",
                "Idle %",
                "Stopped %",
                "Active %",
                "Avg Speed",
                "Avg Moving Speed",
                "Max Speed",
                "Dist / Running Hr",
                "Dist / Motor Hr",
                "Idle / Running Ratio",
            ]
            st.dataframe(daily_display, use_container_width=True)

        with tab2:
            daily_odo = metrics.daily_odometer
            odo_display = daily_odo[["date", "start_odometer", "end_odometer", "avg_odometer", "distance_traveled"]].copy()
            odo_display.columns = ["Date", "Start (km)", "End (km)", "Avg (km)", "Distance (km)"]
            st.dataframe(odo_display, use_container_width=True)

        st.markdown("---")
        st.subheader("Weekly Metrics and Odometer")
        tab3, tab4 = st.tabs(["Weekly Hours & Percentages", "Weekly Odometer"])

        with tab3:
            fig3 = kpi.create_weekly_hours_chart(metrics.weekly_metrics)
            generated_figs.append(fig3)
            st.plotly_chart(fig3, use_container_width=True)

            fig4 = kpi.create_weekly_percentage_chart(metrics.weekly_metrics)
            generated_figs.append(fig4)
            st.plotly_chart(fig4, use_container_width=True)

            weekly_display = metrics.weekly_metrics[
                [
                    "week",
                    "running_hours",
                    "idle_hours",
                    "stopped_hours",
                    "active_hours",
                    "motor_hours",
                    "distance_km",
                    "running_pct",
                    "idle_pct",
                    "stopped_pct",
                    "active_pct",
                    "avg_speed",
                    "avg_moving_speed",
                    "max_speed",
                    "distance_per_running_hour",
                    "distance_per_motor_hour",
                    "idle_to_running_ratio",
                ]
            ].copy()
            weekly_display.columns = [
                "Week",
                "Running Hrs",
                "Idle Hrs",
                "Stopped Hrs",
                "Active Hrs",
                "Motor Hrs",
                "Distance (km)",
                "Running %",
                "Idle %",
                "Stopped %",
                "Active %",
                "Avg Speed",
                "Avg Moving Speed",
                "Max Speed",
                "Dist / Running Hr",
                "Dist / Motor Hr",
                "Idle / Running Ratio",
            ]
            st.dataframe(weekly_display, use_container_width=True)

        with tab4:
            weekly_odo = metrics.weekly_odometer
            weekly_display = weekly_odo[["week", "start_odometer", "end_odometer", "avg_odometer", "distance_traveled"]].copy()
            weekly_display.columns = ["Week", "Start (km)", "End (km)", "Avg (km)", "Distance (km)"]
            st.dataframe(weekly_display, use_container_width=True)

        if generated_figs:
            # Export interactive Plotly charts into a single HTML file.
            report_title = f"kpi_dashboard_{plate_number}_{start_date}_to_{end_date}.html"

            html_body = "".join(
                [
                    pio.to_html(g, full_html=False, include_plotlyjs=("cdn" if i == 0 else False))
                    for i, g in enumerate(generated_figs)
                ]
            )

            # Add a simple title header inside the HTML export for context.
            html_title = (
                "<div style=\""
                "width:100%;"
                "text-align:center;"
                "margin: 10px 0 22px 0;"
                "padding: 18px 18px;"
                "background: linear-gradient(135deg, #0EA5E9 0%, #7C3AED 55%, #F97316 100%);"
                "border-radius: 16px;"
                "box-shadow: 0 10px 30px rgba(0,0,0,0.12);"
                "\">"
                "<div style=\"font-family: 'Segoe UI', Arial, sans-serif; color:white; font-weight:800; "
                "font-size: 28px; line-height:1.2; letter-spacing:0.2px;\">"
                "Vehicle Performance Report"
                "</div>"
                "<div style=\"font-family: 'Segoe UI', Arial, sans-serif; color: rgba(255,255,255,0.95); "
                "font-size: 15px; margin-top: 10px;\">"
                f"<span style=\"font-weight:700;\">Plate Number:</span> {plate_number}"
                " &nbsp; &nbsp; "
                f"<span style=\"font-weight:700;\">Period:</span> {start_date} to {end_date}"
                "<br/>"
                f"<span style=\"font-weight:700;\">TZ:</span> {selected_tz}"
                "</div>"
                "</div>"
            )
            html_body_with_title = html_title + html_body

            st.download_button(
                label="Save KPI Dashboard as Interactive HTML",
                data=html_body_with_title,
                file_name=report_title,
                mime="text/html",
                use_container_width=True,
            )

            st.markdown("### Send KPI HTML by Email")
            st.caption(
                "Recipients receive the HTML file as an attachment and can open it in any browser."
            )

            mail_col1, mail_col2 = st.columns(2)
            with mail_col1:
                to_input = st.text_input(
                    "To (comma separated)",
                    placeholder="person1@company.com, person2@company.com",
                    key="kpi_mail_to",
                )
                cc_input = st.text_input(
                    "CC (optional)",
                    placeholder="manager@company.com",
                    key="kpi_mail_cc",
                )
                from_email = st.text_input(
                    "From Email (shown to recipient)",
                    value="",
                    placeholder="prashanth.nataraj@tld-maini.com",
                    key="kpi_mail_from",
                )
                sender_name = st.text_input(
                    "Sender Name",
                    value="LINKFMS Telemetry",
                    key="kpi_mail_sender_name",
                )

            with mail_col2:
                smtp_host = st.text_input(
                    "SMTP Host",
                    value="smtp.office365.com",
                    key="kpi_mail_smtp_host",
                )
                smtp_port = st.number_input(
                    "SMTP Port",
                    min_value=1,
                    max_value=65535,
                    value=587,
                    key="kpi_mail_smtp_port",
                )
                smtp_username = st.text_input(
                    "SMTP Username",
                    value="",
                    placeholder="Usually same as mailbox username",
                    key="kpi_mail_user",
                )
                smtp_password = st.text_input(
                    "SMTP Password",
                    type="password",
                    key="kpi_mail_password",
                )

            subject = st.text_input(
                "Email Subject",
                value=f"KPI Dashboard Report - {plate_number} ({start_date} to {end_date})",
                key="kpi_mail_subject",
            )
            body_text = st.text_area(
                "Email Body",
                value=(
                    f"Hello,\n\nPlease find attached the KPI dashboard HTML report for {plate_number} "
                    f"from {start_date} to {end_date}.\n\nRegards,\n{sender_name}"
                ),
                key="kpi_mail_body",
            )
            use_tls = st.checkbox("Use STARTTLS", value=True, key="kpi_mail_tls")

            if st.button("Send KPI HTML Email", type="primary", use_container_width=True):
                to_emails = parse_email_list(to_input)
                cc_emails = parse_email_list(cc_input)

                required_values = [
                    smtp_host,
                    smtp_username,
                    smtp_password,
                    from_email,
                    subject,
                ]
                if not to_emails:
                    st.error("Please provide at least one valid recipient email in To.")
                elif not all(required_values):
                    st.error(
                        "Please fill required fields: SMTP host, SMTP username, SMTP password, From email, and Subject."
                    )
                else:
                    with st.spinner("Sending report email..."):
                        send_result = send_html_report_email(
                            smtp_host=smtp_host,
                            smtp_port=int(smtp_port),
                            smtp_username=smtp_username,
                            smtp_password=smtp_password,
                            from_email=from_email,
                            sender_name=sender_name,
                            to_emails=to_emails,
                            cc_emails=cc_emails,
                            subject=subject,
                            body_text=body_text,
                            html_file_name=report_title,
                            html_content=html_body_with_title,
                            use_tls=use_tls,
                        )

                    if send_result.success:
                        st.success(send_result.message)
                    else:
                        st.error(send_result.message)

                    st.info(f"Sender display note: {send_result.sender_identity_hint}")

    elif data_type == "Advanced KPI Insights (TMX Style)":
        if "display_df" not in st.session_state:
            st.info("Fetch telemetry data first, then open Advanced KPI Insights.")
            return

        df = st.session_state["display_df"]
        plate_number = st.session_state.get("tele_plate", "Unknown")
        start_date = st.session_state.get("tele_start", "N/A")
        end_date = st.session_state.get("tele_end", "N/A")
        selected_tz = st.session_state.get("tele_tz", "UTC")

        st.markdown(
            f"Advanced KPI Insights (TMX Style)\n**Plate:** {plate_number} | **Period:** {start_date} to {end_date} | **TZ:** {selected_tz}"
        )
        st.markdown("---")

        kpi = TMXKPIAgent()
        timestamp_col = "dateProcessed" if "dateProcessed" in df.columns else "timestamp"
        ts_min = pd.to_datetime(df[timestamp_col], errors="coerce").min() if timestamp_col in df.columns else None
        ts_max = pd.to_datetime(df[timestamp_col], errors="coerce").max() if timestamp_col in df.columns else None
        advanced_cache_key = (
            str(plate_number),
            str(start_date),
            str(end_date),
            str(selected_tz),
            int(len(df)),
            str(ts_min),
            str(ts_max),
        )

        cached_advanced = st.session_state.get("advanced_kpi_cached_payload")
        cached_key = st.session_state.get("advanced_kpi_cached_key")

        if cached_advanced is not None and cached_key == advanced_cache_key:
            advanced = cached_advanced["advanced"]
            generated_figs = cached_advanced["figs"]
        else:
            advanced = kpi.build_advanced_kpi_bundle(df, timestamp_col=timestamp_col)
            generated_figs = []

            fig_state_pct = kpi.create_engine_state_pct_chart(advanced["state_day"], advanced["state_week"])
            generated_figs.append(fig_state_pct)

            fig_charge = kpi.create_charge_cycles_chart(
                advanced["cycle_detail"], advanced["cycle_day"], advanced["cycle_week"]
            )
            generated_figs.append(fig_charge)

            fig_usage_bar = kpi.create_usage_hours_bar_chart(advanced["day_rollup"], advanced["week_rollup"])
            generated_figs.append(fig_usage_bar)

            fig_odo_bar = kpi.create_odometer_bar_chart(advanced["day_rollup"], advanced["week_rollup"])
            generated_figs.append(fig_odo_bar)

            fig_speed_bar = kpi.create_speed_bar_chart(advanced["day_rollup"], advanced["week_rollup"])
            generated_figs.append(fig_speed_bar)

            fig_idle_bar = kpi.create_idle_hours_bar_chart(advanced["day_rollup"], advanced["week_rollup"])
            generated_figs.append(fig_idle_bar)

            fig_deadman = kpi.create_deadman_switch_chart(advanced["deadman_df"])
            generated_figs.append(fig_deadman)

            fig_soc = kpi.create_soc_chart(advanced["soc_df"])
            generated_figs.append(fig_soc)

            fig_current = kpi.create_battery_current_chart(advanced["battery_current_df"])
            generated_figs.append(fig_current)

            fig_cell = kpi.create_cell_voltage_chart(advanced["cell_voltage_df"])
            generated_figs.append(fig_cell)

            fig_bat_v = kpi.create_battery_voltage_chart(advanced["battery_voltage_df"])
            generated_figs.append(fig_bat_v)

            fig_raw = kpi.create_raw_telemetry_line_chart(advanced["raw_df"])
            generated_figs.append(fig_raw)

            st.session_state["advanced_kpi_cached_key"] = advanced_cache_key
            st.session_state["advanced_kpi_cached_payload"] = {
                "advanced": advanced,
                "figs": generated_figs,
            }

        adv_kpis = advanced["kpis"]

        adv_col1, adv_col2, adv_col3, adv_col4 = st.columns(4)
        with adv_col1:
            st.metric("Avg Charge Cycles / Day", f"{adv_kpis['avg_charge_cycles_day']:.2f}")
        with adv_col2:
            st.metric("Avg Charge Cycles / Week", f"{adv_kpis['avg_charge_cycles_week']:.2f}")
        with adv_col3:
            st.metric("Avg Running % / Day", f"{adv_kpis['avg_running_pct_day']:.1f}%")
        with adv_col4:
            st.metric("Avg Idle % / Day", f"{adv_kpis['avg_idle_pct_day']:.1f}%")

        adv_col5, adv_col6, adv_col7, adv_col8 = st.columns(4)
        with adv_col5:
            st.metric("Odometer / Day", f"{adv_kpis['avg_odometer_day_km']:.2f} km")
        with adv_col6:
            st.metric("Odometer / Week", f"{adv_kpis['avg_odometer_week_km']:.2f} km")
        with adv_col7:
            st.metric("Motor Hour / Day", f"{adv_kpis['avg_motor_hour_day']:.2f} h")
        with adv_col8:
            st.metric("Motor Hour / Week", f"{adv_kpis['avg_motor_hour_week']:.2f} h")

        adv_col9, adv_col10, adv_col11, adv_col12 = st.columns(4)
        with adv_col9:
            st.metric("Avg Speed / Day", f"{adv_kpis['avg_speed_day']:.2f}")
        with adv_col10:
            st.metric("Avg Speed / Week", f"{adv_kpis['avg_speed_week']:.2f}")
        with adv_col11:
            st.metric("Idle Hours / Day", f"{adv_kpis['avg_idle_hours_day']:.2f} h")
        with adv_col12:
            st.metric("Idle Hours / Week", f"{adv_kpis['avg_idle_hours_week']:.2f} h")

        for fig in generated_figs:
            st.plotly_chart(fig, use_container_width=True)

        if generated_figs:
            report_title = f"advanced_kpi_dashboard_{plate_number}_{start_date}_to_{end_date}.html"
            html_body = "".join(
                [
                    pio.to_html(g, full_html=False, include_plotlyjs=("cdn" if i == 0 else False))
                    for i, g in enumerate(generated_figs)
                ]
            )

            adv_cards = [
                ("Avg Charge Cycles / Day", f"{adv_kpis['avg_charge_cycles_day']:.2f}"),
                ("Avg Charge Cycles / Week", f"{adv_kpis['avg_charge_cycles_week']:.2f}"),
                ("Avg Running % / Day", f"{adv_kpis['avg_running_pct_day']:.1f}%"),
                ("Avg Idle % / Day", f"{adv_kpis['avg_idle_pct_day']:.1f}%"),
                ("Odometer / Day", f"{adv_kpis['avg_odometer_day_km']:.2f} km"),
                ("Odometer / Week", f"{adv_kpis['avg_odometer_week_km']:.2f} km"),
                ("Motor Hour / Day", f"{adv_kpis['avg_motor_hour_day']:.2f} h"),
                ("Motor Hour / Week", f"{adv_kpis['avg_motor_hour_week']:.2f} h"),
                ("Avg Speed / Day", f"{adv_kpis['avg_speed_day']:.2f}"),
                ("Avg Speed / Week", f"{adv_kpis['avg_speed_week']:.2f}"),
                ("Idle Hours / Day", f"{adv_kpis['avg_idle_hours_day']:.2f} h"),
                ("Idle Hours / Week", f"{adv_kpis['avg_idle_hours_week']:.2f} h"),
            ]
            adv_cards_html = "".join(
                [
                    (
                        "<div style='background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:12px 14px;box-shadow:0 2px 8px rgba(0,0,0,0.05);'>"
                        f"<div style='font-size:12px;color:#6b7280;margin-bottom:6px;'>{label}</div>"
                        f"<div style='font-size:22px;font-weight:700;color:#0f172a;'>{value}</div>"
                        "</div>"
                    )
                    for label, value in adv_cards
                ]
            )

            html_with_cards = (
                "<!DOCTYPE html><html lang='en'><head><meta charset='utf-8'/>"
                "<meta name='viewport' content='width=device-width, initial-scale=1'/>"
                "<title>Advanced KPI Dashboard</title></head>"
                "<body style='margin:0;padding:16px;background:#f8fafc;color:#0f172a;font-family:Segoe UI,Arial,sans-serif;'>"
                "<div style='width:100%;text-align:center;margin:6px 0 18px 0;padding:18px;background:linear-gradient(135deg,#0EA5E9 0%,#7C3AED 55%,#F97316 100%);border-radius:14px;color:white;'>"
                "<div style='font-size:26px;font-weight:800;'>Advanced KPI Insights (TMX Style)</div>"
                f"<div style='font-size:14px;margin-top:8px;'><b>Plate:</b> {plate_number} &nbsp; <b>Period:</b> {start_date} to {end_date} &nbsp; <b>TZ:</b> {selected_tz}</div>"
                "</div>"
                "<div style='display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:12px;margin-bottom:18px;'>"
                f"{adv_cards_html}"
                "</div>"
                "<div style='display:grid;gap:16px;'>"
                f"{html_body}"
                "</div></body></html>"
            )

            st.download_button(
                label="Save Advanced KPI Dashboard as Interactive HTML",
                data=html_with_cards,
                file_name=report_title,
                mime="text/html",
                use_container_width=True,
            )

    else:  # AI Assistant
        if "display_df" not in st.session_state:
            st.info("Fetch telemetry data first, then use AI Assistant.")
            return

        ai_df = st.session_state["display_df"]
        ai_plate = st.session_state.get("tele_plate", plate_number or "Unknown")
        ai_start = str(st.session_state.get("tele_start", start_date))
        ai_end = str(st.session_state.get("tele_end", end_date))
        ai_tz = st.session_state.get("tele_tz", "UTC")

        st.subheader("AI Assistant (Fast Local Model)")
        st.caption("Optimized for short and fast answers using local free models via Ollama.")

        col_ai_1, col_ai_2, col_ai_3 = st.columns(3)
        with col_ai_1:
            ollama_model = st.selectbox(
                "Model",
                options=["qwen2.5:3b", "llama3.2:1b", "llama3.2:3b"],
                index=0,
            )
        with col_ai_2:
            ai_timeout = st.slider("Timeout (seconds)", min_value=5, max_value=30, value=12)
        with col_ai_3:
            ai_max_tokens = st.slider("Max Response Tokens", min_value=60, max_value=260, value=140, step=20)

        ollama_host = st.text_input(
            "Ollama Host URL",
            value="http://localhost:11434",
            help="If Ollama runs on another machine, provide its URL (example: http://10.11.192.94:11434).",
        )

        sample_size = st.slider("Context Sample Rows", min_value=4, max_value=20, value=10)

        sample_df = ai_df.head(sample_size).copy()
        sample_records = sample_df.where(sample_df.notna(), None).to_dict(orient="records")
        context_json = build_compact_context(
            plate_number=ai_plate,
            start_date=ai_start,
            end_date=ai_end,
            timezone_label=ai_tz,
            total_rows=len(ai_df),
            columns=list(ai_df.columns),
            sample_rows=sample_records,
        )

        with st.expander("AI Context Preview"):
            st.code(context_json, language="json")

        default_question = (
            "Give a quick summary of vehicle behavior and mention any obvious risk indicators."
        )
        question = st.text_area("Ask a question", value=default_question, height=120)

        if st.button("Ask AI", type="primary", use_container_width=True):
            if not question.strip():
                st.error("Please enter a question.")
            else:
                with st.spinner("Generating fast response..."):
                    ai_result = ask_ollama_fast(
                        question=question.strip(),
                        context_json=context_json,
                        model=ollama_model,
                        host=ollama_host.strip(),
                        timeout_seconds=int(ai_timeout),
                        max_tokens=int(ai_max_tokens),
                    )

                if ai_result.success:
                    st.success("AI response ready")
                    st.caption(f"Model latency: {ai_result.latency_ms} ms")
                    st.write(ai_result.answer)
                else:
                    st.error(ai_result.error or "AI request failed.")
                    st.info(
                        "Quick setup: install Ollama, run it, then pull a model (example: ollama pull qwen2.5:3b)."
                    )


# "?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?
# ENTRY POINT
# "?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?"?
if __name__ == "__main__":
    if st.session_state.authenticated:
        show_main_app()
    else:
        render_login_page(title="🌐 TLD/XOPS VISUALIZATION TOOL")





