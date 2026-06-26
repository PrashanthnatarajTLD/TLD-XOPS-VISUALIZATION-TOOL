"""Utilities for handling file I/O operations."""

import pandas as pd
from pathlib import Path
from typing import Tuple, Optional

from openpyxl import load_workbook
from openpyxl.styles import Font


class FileHandler:
    """Handles file reading and writing operations."""
    
    SUPPORTED_FORMATS = ['.csv', '.xlsx', '.xls']

    @staticmethod
    def _apply_excel_bold_styling(filepath: Path) -> None:
        """
        Apply basic Excel styling so that header/index-like values are bold.

        This is intentionally generic:
        - Row 1 (headers) are bold for every sheet.
        - Column A (index/labels column in pandas outputs like describe()) are bold
          when the cell value is a string.
        """
        wb = load_workbook(filepath)
        bold_font = Font(bold=True)

        for ws in wb.worksheets:
            # Header row
            for cell in ws[1]:
                cell.font = bold_font

            # Index/label column (commonly column A)
            max_row = ws.max_row or 1
            for r in range(2, max_row + 1):
                cell = ws.cell(row=r, column=1)
                if isinstance(cell.value, str) and cell.value.strip() != "":
                    cell.font = bold_font

        wb.save(filepath)
    
    @staticmethod
    def read_file(filepath: str) -> pd.DataFrame:
        """
        Read data from file (CSV, Excel).
        
        Args:
            filepath: Path to the file
            
        Returns:
            DataFrame with the loaded data
            
        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file format not supported
        """
        path = Path(filepath)
        
        if not path.exists():
            raise FileNotFoundError(f"File not found: {filepath}")
        
        file_ext = path.suffix.lower()
        
        if file_ext == '.csv':
            return pd.read_csv(filepath)
        elif file_ext in ['.xlsx', '.xls']:
            return pd.read_excel(filepath)
        else:
            raise ValueError(f"Unsupported file format: {file_ext}. Supported: {FileHandler.SUPPORTED_FORMATS}")
    
    @staticmethod
    def read_excel_sheets(filepath: str) -> dict:
        """
        Read all sheets from an Excel file.
        
        Args:
            filepath: Path to the Excel file
            
        Returns:
            Dictionary with sheet names as keys and DataFrames as values
        """
        path = Path(filepath)
        
        if not path.exists():
            raise FileNotFoundError(f"File not found: {filepath}")
        
        excel_file = pd.ExcelFile(filepath)
        sheets = {}
        
        for sheet_name in excel_file.sheet_names:
            sheets[sheet_name] = pd.read_excel(filepath, sheet_name=sheet_name)
        
        return sheets
    
    @staticmethod
    def get_sheet_names(filepath: str) -> list:
        """Get list of sheet names from Excel file."""
        path = Path(filepath)
        
        if not path.exists():
            raise FileNotFoundError(f"File not found: {filepath}")
        
        excel_file = pd.ExcelFile(filepath)
        return excel_file.sheet_names
    
    @staticmethod
    def save_csv(dataframe: pd.DataFrame, filepath: str) -> None:
        """Save DataFrame to CSV."""
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)
        dataframe.to_csv(filepath, index=False)
    
    @staticmethod
    def save_excel(dataframe: pd.DataFrame, filepath: str) -> None:
        """Save DataFrame to Excel (atomic write to avoid partially-written XLSX)."""
        dest_path = Path(filepath)
        dest_path.parent.mkdir(parents=True, exist_ok=True)

        # Write to temp file in the same directory, then replace atomically.
        tmp_path = dest_path.with_suffix(dest_path.suffix + ".tmp.xlsx")

        # Use openpyxl explicitly for stable .xlsx generation.
        dataframe.to_excel(tmp_path, index=False, engine="openpyxl")

        # Apply styling (bold header + bold index/labels column).
        FileHandler._apply_excel_bold_styling(tmp_path)

        # Replace the destination with the temp file.
        tmp_path.replace(dest_path)
    
    @staticmethod
    def save_multiple_sheets(sheets: dict, filepath: str) -> None:
        """
        Save multiple DataFrames to different sheets in Excel (atomic write).
        
        Args:
            sheets: Dictionary with sheet names as keys and DataFrames as values
            filepath: Output Excel file path
        """
        dest_path = Path(filepath)
        dest_path.parent.mkdir(parents=True, exist_ok=True)

        tmp_path = dest_path.with_suffix(dest_path.suffix + ".tmp.xlsx")

        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Apply styling (bold header + bold index/labels column) across all sheets.
        FileHandler._apply_excel_bold_styling(tmp_path)

        tmp_path.replace(dest_path)
